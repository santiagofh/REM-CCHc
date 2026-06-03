from __future__ import annotations

import argparse
import csv
import json
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
OUTPUT_DIR = BASE_DIR / "output"
DEFAULT_JSON_PATH = DATA_DIR / "diccionario_rem_chcc_2025.json"
DEFAULT_CSV_PATH = Path(
    r"D:\DATA\REM\REM_2025\Datos\SerieA2025.csv"
)
EXTERNAL_ESTABLISHMENTS_DIR = Path(
    r"C:\Users\fariass\OneDrive - SUBSECRETARIA DE SALUD PUBLICA\Escritorio\DATA\ESTABLECIMIENTOS"
)
DEFAULT_WORKBOOK_PATH = OUTPUT_DIR / "reporte_indicadores_chcc_2025.xlsx"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Genera un Excel consolidado por establecimiento para los indicadores A2, A4 y H2."
    )
    parser.add_argument("--json-path", type=Path, default=DEFAULT_JSON_PATH)
    parser.add_argument("--csv-path", type=Path, default=DEFAULT_CSV_PATH)
    parser.add_argument("--workbook-path", type=Path, default=DEFAULT_WORKBOOK_PATH)
    parser.add_argument("--ano", default="2025")
    parser.add_argument("--mes")
    parser.add_argument("--id-servicio")
    parser.add_argument("--id-region")
    parser.add_argument("--id-comuna")
    parser.add_argument("--id-establecimiento")
    return parser.parse_args()


def load_dictionary(json_path: Path) -> dict:
    with json_path.open("r", encoding="utf-8") as file:
        return json.load(file)


def safe_int(value: str | None) -> int:
    if value is None:
        return 0
    text = str(value).strip()
    if not text:
        return 0
    return int(float(text))


def normalize_spaces(text: str) -> str:
    return " ".join(text.split())


def find_codes_by_meaning(section_codes: dict[str, str], exact_meanings: set[str]) -> set[str]:
    expected = {normalize_spaces(item) for item in exact_meanings}
    found = {
        code
        for code, meaning in section_codes.items()
        if normalize_spaces(meaning) in expected
    }
    missing = expected - {normalize_spaces(meaning) for code, meaning in section_codes.items() if code in found}
    if missing:
        raise ValueError(f"No se encontraron significados esperados en el JSON: {', '.join(sorted(missing))}")
    return found


def find_denominator_codes_h2(section_codes: dict[str, str]) -> set[str]:
    included_prefixes = (
        "Vaginal",
        "Instrumental",
        "Cesárea Electiva",
        "Cesárea Urgencia",
        "Parto prehospitalario",
        "Partos fuera de la red de salud",
        "Parto en domicilio",
    )
    codes = {
        code
        for code, meaning in section_codes.items()
        if normalize_spaces(meaning).startswith(included_prefixes)
    }
    if not codes:
        raise ValueError("No se pudieron identificar los codigos del denominador para H2.")
    return codes


def resolve_establishments_path() -> Path:
    local_files = sorted(DATA_DIR.glob("establecimientos_*.csv"))
    if local_files:
        return local_files[-1]

    external_files = sorted(EXTERNAL_ESTABLISHMENTS_DIR.glob("establecimientos_*.csv"))
    if external_files:
        return external_files[-1]

    raise FileNotFoundError("No se encontro un archivo de establecimientos en data/ ni en DATA/ESTABLECIMIENTOS.")


def load_establishments_map(path: Path) -> dict[str, dict[str, str]]:
    mapping: dict[str, dict[str, str]] = {}
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file, delimiter=";")
        for row in reader:
            code = str(row.get("EstablecimientoCodigo", "")).strip()
            if not code or code in mapping:
                continue
            mapping[code] = {
                "Region": row.get("RegionGlosa", "").strip(),
                "Servicio": row.get("SeremiSaludGlosa_ServicioDeSaludGlosa", "").strip(),
                "Comuna": row.get("ComunaGlosa", "").strip(),
                "Establecimiento": row.get("EstablecimientoGlosa", "").strip(),
                "RegionCodigo": str(row.get("RegionCodigo", "")).strip(),
                "ServicioCodigo": str(row.get("SeremiSaludCodigo_ServicioDeSaludCodigo", "")).strip(),
                "ComunaCodigo": str(row.get("ComunaCodigo", "")).strip(),
            }
    return mapping


def matches_filters(row: dict, args: argparse.Namespace) -> bool:
    filters = {
        "Ano": args.ano,
        "Mes": args.mes,
        "IdServicio": args.id_servicio,
        "IdRegion": args.id_region,
        "IdComuna": args.id_comuna,
        "IdEstablecimiento": args.id_establecimiento,
    }
    for field, expected in filters.items():
        if expected is None:
            continue
        if str(row.get(field, "")).strip() != str(expected):
            return False
    return True


def build_indicator_definitions(dictionary: dict) -> dict:
    section_a27 = dictionary["REM A27"]["SECCIÓN A: PERSONAS QUE INGRESAN A EDUCACIÓN GRUPAL SEGÚN ÁREAS TEMÁTICAS Y EDAD"]
    section_a05_a = dictionary["REM A05"]["SECCIÓN A: INGRESOS DE GESTANTES A PROGRAMA PRENATAL"]
    section_a05_e = dictionary["REM A05"]["SECCIÓN E: INGRESOS A CONTROL DE SALUD DE RECIÉN NACIDOS"]
    section_a01 = dictionary["REM A01"]["SECCIÓN A: CONTROLES DE SALUD SEXUAL Y REPRODUCTIVA"]
    section_h2_general = dictionary["REM A024"]["SECCIÓN A: INFORMACIÓN GENERAL DE PARTOS"]
    section_h2_vaginal = dictionary["REM A024"]["SECCION A.1: PARTOS VAGINALES *"]
    section_h2_cesarea = dictionary["REM A024"]["SECCION A.2: CESÁREAS (RESPONSABILIDAD DEL MÉDICO JEFE DEL SERVICIO DE OBSTETRICIA)"]

    defs = {
        "A2": {
            "titulo": "A2 - Desarrollo Prenatal",
            "nombre": "Porcentaje de gestantes que ingresan a educación grupal presencial o remota en APS",
            "meta": 0.80,
            "ponderacion": 0.15,
            "relevant_codes": {"27500110", "01080008"},
            "apply_row": lambda row, agg: apply_a2_row(row, agg),
        },
        "A4": {
            "titulo": "A4 - Control de salud del niño y niña para el desarrollo integral",
            "nombre": "Porcentaje de controles de salud entregados a díadas dentro de los 10 días de vida",
            "meta": 0.70,
            "ponderacion": 0.15,
            "relevant_codes": {"01110106", "01110107", "05225100"},
            "apply_row": lambda row, agg: apply_a4_row(row, agg),
        },
        "H2": {
            "titulo": "H2 - Atención personalizada del parto",
            "nombre": "Porcentaje de mujeres con acompañamiento durante el preparto y parto",
            "meta": None,
            "ponderacion": None,
            "relevant_codes": find_denominator_codes_h2(section_h2_general["codigos"])
            | find_codes_by_meaning(section_h2_vaginal["codigos"], {"Acompañamiento - Durante el trabajo de parto"})
            | find_codes_by_meaning(section_h2_cesarea["codigos"], {"Acompañamiento durante la cesárea"}),
            "apply_row": lambda row, agg: apply_h2_row(row, agg),
        },
    }
    return defs


def apply_a2_row(row: dict, agg: dict) -> None:
    code = row["CodigoPrestacion"].strip()
    if code == "27500110":
        agg["numerador"] += safe_int(row.get("Col22"))
    elif code == "01080008":
        agg["denominador"] += safe_int(row.get("Col01"))


def apply_a4_row(row: dict, agg: dict) -> None:
    code = row["CodigoPrestacion"].strip()
    if code in {"01110106", "01110107"}:
        agg["numerador"] += safe_int(row.get("Col01"))
    elif code == "05225100":
        agg["denominador"] += safe_int(row.get("Col01"))


def apply_h2_row(row: dict, agg: dict) -> None:
    code = row["CodigoPrestacion"].strip()
    if code in {"01030100", "01030300", "24090700", "29101714", "29101715", "29101716", "29101717", "29101718"}:
        agg["denominador"] += safe_int(row.get("Col01"))
    elif code == "29101728":
        agg["numerador"] += safe_int(row.get("Col01"))
    elif code == "29101742":
        agg["numerador"] += safe_int(row.get("Col01")) + safe_int(row.get("Col02"))


def process_csv(csv_path: Path, args: argparse.Namespace, indicator_defs: dict) -> dict[str, dict[tuple, dict]]:
    results: dict[str, dict[tuple, dict]] = {
        indicator: defaultdict(
            lambda: {
                "Mes": 0,
                "RegionCodigo": "",
                "ServicioCodigo": "",
                "ComunaCodigo": "",
                "numerador": 0,
                "denominador": 0,
            }
        )
        for indicator in indicator_defs
    }

    all_codes = set().union(*(indicator_defs[ind]["relevant_codes"] for ind in indicator_defs))

    with csv_path.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file, delimiter=";")
        for row in reader:
            if not matches_filters(row, args):
                continue

            code = str(row["CodigoPrestacion"]).strip()
            if code not in all_codes:
                continue

            est_code = str(row["IdEstablecimiento"]).strip()
            mes = str(row.get("Mes", "")).strip()
            for indicator, config in indicator_defs.items():
                if code not in config["relevant_codes"]:
                    continue
                agg = results[indicator][(est_code, mes)]
                if not agg["Mes"]:
                    agg["Mes"] = int(mes)
                if not agg["RegionCodigo"]:
                    agg["RegionCodigo"] = str(row.get("IdRegion", "")).strip()
                    agg["ServicioCodigo"] = str(row.get("IdServicio", "")).strip()
                    agg["ComunaCodigo"] = str(row.get("IdComuna", "")).strip()
                config["apply_row"](row, agg)

    return results


def build_rows_for_indicator(
    indicator: str,
    raw_rows: dict[tuple, dict],
    establishments_map: dict[str, dict[str, str]],
) -> list[dict]:
    rows = []
    for (est_code, mes), values in raw_rows.items():
        if values["numerador"] == 0 and values["denominador"] == 0:
            continue

        est_info = establishments_map.get(est_code, {})
        denominator = values["denominador"]
        percentage_decimal = (values["numerador"] / denominator) if denominator else None

        rows.append(
            {
                "Region": est_info.get("Region") or f"Código {values['RegionCodigo']}",
                "Servicio de salud": est_info.get("Servicio") or f"Código {values['ServicioCodigo']}",
                "Comuna": est_info.get("Comuna") or f"Código {values['ComunaCodigo']}",
                "Establecimiento": est_info.get("Establecimiento") or f"Establecimiento {est_code}",
                "CodigoEstablecimiento": est_code,
                "Mes": values["Mes"],
                "Numerador": values["numerador"],
                "Denominador": denominator,
                "PorcentajeCumplimiento": percentage_decimal,
            }
        )

    rows.sort(
        key=lambda item: (
            item["Region"],
            item["Servicio de salud"],
            item["Comuna"],
            item["Establecimiento"],
            item["Mes"],
        )
    )
    return rows


def write_indicator_csv(path: Path, rows: list[dict]) -> None:
    fieldnames = [
        "Region",
        "Servicio de salud",
        "Comuna",
        "Establecimiento",
        "CodigoEstablecimiento",
        "Mes",
        "Numerador",
        "Denominador",
        "PorcentajeCumplimiento",
    ]
    with path.open("w", encoding="utf-8", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames, delimiter=";")
        writer.writeheader()
        for row in rows:
            serializable = dict(row)
            if serializable["PorcentajeCumplimiento"] is None:
                serializable["PorcentajeCumplimiento"] = ""
            else:
                serializable["PorcentajeCumplimiento"] = round(serializable["PorcentajeCumplimiento"] * 100, 2)
            writer.writerow(serializable)


def style_sheet(ws, rows_count: int, headers_count: int) -> None:
    ws.sheet_view.showGridLines = False
    last_col = chr(ord("A") + headers_count - 1)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True, color="000000")
    thin = Side(style="thin", color="D1D5DB")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows(min_row=2, max_row=1 + rows_count, min_col=1, max_col=headers_count):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{last_col}{1 + rows_count}"

    # Usamos autofiltro simple en vez de tablas de Excel para evitar reparaciones al abrir.


def build_summary_rows(all_rows: dict[str, list[dict]], level: str) -> list[dict]:
    if level == "SS":
        id_fields = ["Region", "Servicio de salud"]
    elif level == "Comuna":
        id_fields = ["Region", "Servicio de salud", "Comuna"]
    elif level == "Establecimiento":
        id_fields = ["Region", "Servicio de salud", "Comuna", "Establecimiento"]
    else:
        raise ValueError(f"Nivel no soportado: {level}")

    grouped: dict[tuple, dict] = {}
    for indicator, rows in all_rows.items():
        for row in rows:
            key = tuple(row.get(field, "") for field in id_fields)
            current = grouped.setdefault(
                key,
                {field: row.get(field, "") for field in id_fields},
            )
            current[f"{indicator}_Numerador"] = current.get(f"{indicator}_Numerador", 0) + row["Numerador"]
            current[f"{indicator}_Denominador"] = current.get(f"{indicator}_Denominador", 0) + row["Denominador"]

    summary_rows = []
    for key in sorted(grouped.keys()):
        current = grouped[key]
        for indicator in ["A2", "A4", "H2"]:
            numerator = current.get(f"{indicator}_Numerador", 0)
            denominator = current.get(f"{indicator}_Denominador", 0)
            current[f"{indicator}_Porcentaje"] = (numerator / denominator) if denominator else None
        summary_rows.append(current)

    return summary_rows


def set_column_widths(ws, widths: list[int]) -> None:
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(ord("A") + idx - 1)].width = width


def format_percentage_columns(ws, row_count: int, percentage_columns: list[int], number_columns: list[int]) -> None:
    for row_idx in range(2, 2 + row_count):
        for col_idx in number_columns:
            ws.cell(row=row_idx, column=col_idx).number_format = "#,##0"
        for col_idx in percentage_columns:
            ws.cell(row=row_idx, column=col_idx).number_format = "0.00%"


def add_summary_sheet(wb: Workbook, sheet_name: str, rows: list[dict]) -> None:
    ws = wb.create_sheet(title=sheet_name)

    if sheet_name == "SS":
        id_fields = ["Region", "Servicio de salud"]
        widths = [24, 30, 14, 14, 16, 14, 14, 16, 14, 14, 16]
    elif sheet_name == "Comuna":
        id_fields = ["Region", "Servicio de salud", "Comuna"]
        widths = [24, 30, 24, 14, 14, 16, 14, 14, 16, 14, 14, 16]
    else:
        id_fields = ["Region", "Servicio de salud", "Comuna", "Establecimiento"]
        widths = [24, 30, 24, 42, 14, 14, 16, 14, 14, 16, 14, 14, 16]

    headers = id_fields + [
        "A2 Numerador",
        "A2 Denominador",
        "A2 %",
        "A4 Numerador",
        "A4 Denominador",
        "A4 %",
        "H2 Numerador",
        "H2 Denominador",
        "H2 %",
    ]

    ws.append(headers)

    for item in rows:
        ws.append(
            [item.get(field, "") for field in id_fields]
            + [
                item.get("A2_Numerador", 0),
                item.get("A2_Denominador", 0),
                item.get("A2_Porcentaje"),
                item.get("A4_Numerador", 0),
                item.get("A4_Denominador", 0),
                item.get("A4_Porcentaje"),
                item.get("H2_Numerador", 0),
                item.get("H2_Denominador", 0),
                item.get("H2_Porcentaje"),
            ]
        )

    style_sheet(ws, len(rows), len(headers))
    set_column_widths(ws, widths)
    first_metric_col = len(id_fields) + 1
    number_columns = [first_metric_col, first_metric_col + 1, first_metric_col + 3, first_metric_col + 4, first_metric_col + 6, first_metric_col + 7]
    percentage_columns = [first_metric_col + 2, first_metric_col + 5, first_metric_col + 8]
    format_percentage_columns(ws, len(rows), percentage_columns, number_columns)


def add_detail_sheet(wb: Workbook, indicator: str, rows: list[dict], config: dict) -> None:
    ws = wb.create_sheet(title=indicator)
    headers = [
        "Región",
        "Servicio de salud",
        "Comuna",
        "Establecimiento",
        "Mes",
        "Numerador",
        "Denominador",
        "Porcentaje de cumplimiento",
    ]
    ws.append(headers)

    for item in rows:
        ws.append(
            [
                item["Region"],
                item["Servicio de salud"],
                item["Comuna"],
                item["Establecimiento"],
                item["Mes"],
                item["Numerador"],
                item["Denominador"],
                item["PorcentajeCumplimiento"],
            ]
        )

    style_sheet(ws, len(rows), len(headers))
    set_column_widths(ws, [24, 30, 24, 42, 8, 14, 14, 18])
    format_percentage_columns(ws, len(rows), [8], [6, 7])


def create_workbook(workbook_path: Path, all_rows: dict[str, list[dict]], indicator_defs: dict) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    for level in ["SS", "Comuna", "Establecimiento"]:
        summary_rows = build_summary_rows(all_rows, level)
        add_summary_sheet(wb, level, summary_rows)

    for indicator in ["A2", "A4", "H2"]:
        add_detail_sheet(wb, indicator, all_rows[indicator], indicator_defs[indicator])

    wb.save(workbook_path)


def main() -> None:
    args = parse_args()
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    dictionary = load_dictionary(args.json_path)
    indicator_defs = build_indicator_definitions(dictionary)
    establishments_path = resolve_establishments_path()
    establishments_map = load_establishments_map(establishments_path)

    raw_results = process_csv(args.csv_path, args, indicator_defs)

    all_rows = {}
    for indicator in indicator_defs:
        rows = build_rows_for_indicator(indicator, raw_results[indicator], establishments_map)
        all_rows[indicator] = rows
        write_indicator_csv(OUTPUT_DIR / f"resumen_{indicator.lower()}_por_establecimiento.csv", rows)

    create_workbook(args.workbook_path, all_rows, indicator_defs)

    print(f"Establecimientos usados: {establishments_path}")
    print(f"Excel generado: {args.workbook_path}")
    for indicator in ["A2", "A4", "H2"]:
        print(f"{indicator}: {len(all_rows[indicator])} establecimientos")


if __name__ == "__main__":
    main()
